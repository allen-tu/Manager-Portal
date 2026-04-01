#!/usr/bin/env python3
"""
銷售案 Forecast 分析系統 — 後端服務
版本：v1.1.0
執行方式：雙擊 啟動.bat (Windows) 或 啟動.command (Mac)
"""
__version__ = '1.1.0'
import os, json, sys, webbrowser, threading, time, io
from datetime import datetime
from flask import Flask, request, jsonify, send_file, abort

app      = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, 'forecast.db')
HTML_PATH= os.path.join(BASE_DIR, '銷售案分析系統.html')

# ── Database ─────────────────────────────────────────────────────────────────

def get_db():
    import sqlite3
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            );
            INSERT OR IGNORE INTO settings VALUES ('max_history', '12');

            CREATE TABLE IF NOT EXISTS uploads (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                filename     TEXT,
                bu_names     TEXT,
                uploaded_at  TEXT,
                case_count   INTEGER,
                conf_summary TEXT
            );

            CREATE TABLE IF NOT EXISTS cases (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                upload_id  INTEGER REFERENCES uploads(id) ON DELETE CASCADE,
                case_no    TEXT,
                cust_name  TEXT,
                case_name  TEXT,
                rep        TEXT,
                bu         TEXT,
                close_date TEXT,
                dept       TEXT DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS case_items (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                case_id   INTEGER REFERENCES cases(id) ON DELETE CASCADE,
                conf      TEXT,
                item_name TEXT,
                amount    REAL
            );

            CREATE TABLE IF NOT EXISTS quotas (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                year       INTEGER,
                dim_type   TEXT,
                dim_name   TEXT,
                quota      REAL,
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS orders (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                year       INTEGER,
                dim_type   TEXT,
                dim_name   TEXT,
                amount     REAL,
                source     TEXT DEFAULT 'manual',
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS order_uploads (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                filename    TEXT,
                uploaded_at TEXT,
                year        INTEGER,
                row_count   INTEGER
            );

            CREATE TABLE IF NOT EXISTS opp_tracking (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                case_no      TEXT NOT NULL,
                case_name    TEXT DEFAULT '',
                cust_name    TEXT DEFAULT '',
                note         TEXT NOT NULL,
                is_done      INTEGER DEFAULT 0,
                done_note    TEXT DEFAULT '',
                done_at      TEXT DEFAULT '',
                created_at   TEXT
            );

            CREATE TABLE IF NOT EXISTS order_records (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                upload_id        INTEGER REFERENCES order_uploads(id) ON DELETE CASCADE,
                contract_no      TEXT,
                contract_no_bu   TEXT,
                contract_name    TEXT,
                cust_name        TEXT,
                register_date    TEXT,
                register_year    INTEGER,
                register_month   INTEGER,
                amount_pretax    REAL,
                biz_bu           TEXT,
                biz_dept         TEXT,
                biz_rep          TEXT,
                support_bu       TEXT,
                support_dept     TEXT,
                transaction_type TEXT,
                contract_type    TEXT,
                industry         TEXT,
                industry_class   TEXT,
                product_line     TEXT,
                biz_line         TEXT,
                vendor           TEXT,
                product_name     TEXT,
                item_seq         TEXT,
                case_no          TEXT
            );
        """)
        conn.commit()
        # ── 欄位遷移：舊資料庫若缺少欄位則自動補上 ──
        try:
            conn.execute("ALTER TABLE cases ADD COLUMN dept TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass  # 欄位已存在，忽略
        try:
            conn.execute("ALTER TABLE cases ADD COLUMN support_bu TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE cases ADD COLUMN support_dept TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        # case_items 層級的支援欄位（正確做法：支援部門屬於分項，非案件整體）
        try:
            conn.execute("ALTER TABLE case_items ADD COLUMN support_bu TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE case_items ADD COLUMN support_dept TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE orders ADD COLUMN source TEXT DEFAULT 'manual'")
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE order_records ADD COLUMN biz_rep TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE opp_tracking ADD COLUMN done_note TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE opp_tracking ADD COLUMN done_at TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass

# ── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_file(HTML_PATH)

@app.route('/api/ping')
def ping():
    return jsonify({'ok': True})

# ── Settings ──────────────────────────────────────────────────────────────────
@app.route('/api/settings', methods=['GET', 'POST'])
def settings_api():
    with get_db() as conn:
        if request.method == 'POST':
            data = request.get_json() or {}
            for key, value in data.items():
                conn.execute("INSERT OR REPLACE INTO settings VALUES (?,?)", (key, str(value)))
            conn.commit()
            return jsonify({'ok': True})
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
        return jsonify({r['key']: r['value'] for r in rows})

# ── Upload (save parsed data from frontend) ───────────────────────────────────
@app.route('/api/upload', methods=['POST'])
def upload_api():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400

    raw      = data.get('raw', {})
    bu_order = data.get('buOrder', [])
    filename = data.get('filename', '未知檔案')

    # Compute summary totals per conf
    conf_summary = {}
    case_count   = 0
    for bu, cases in raw.items():
        case_count += len(cases)
        for cas in cases:
            for conf, items in (cas.get('conf') or {}).items():
                s = sum(i.get('amt', 0) for i in items)
                conf_summary[conf] = conf_summary.get(conf, 0) + s

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    with get_db() as conn:
        row = conn.execute("SELECT value FROM settings WHERE key='max_history'").fetchone()
        max_history = int(row['value']) if row else 12

        cur = conn.execute(
            "INSERT INTO uploads (filename, bu_names, uploaded_at, case_count, conf_summary) VALUES (?,?,?,?,?)",
            (filename, ','.join(bu_order), now, case_count,
             json.dumps(conf_summary, ensure_ascii=False))
        )
        upload_id = cur.lastrowid

        for bu, cases in raw.items():
            for cas in cases:
                case_cur = conn.execute(
                    "INSERT INTO cases (upload_id,case_no,cust_name,case_name,rep,bu,close_date,dept,support_bu,support_dept) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (upload_id, cas.get('no',''), cas.get('cust',''), cas.get('case',''),
                     cas.get('rep',''), bu, cas.get('date',''), cas.get('dept',''),
                     cas.get('supportBu',''), cas.get('supportDept',''))
                )
                case_id = case_cur.lastrowid
                for conf, items in (cas.get('conf') or {}).items():
                    for item in items:
                        conn.execute(
                            "INSERT INTO case_items (case_id,conf,item_name,amount,support_bu,support_dept) VALUES (?,?,?,?,?,?)",
                            (case_id, conf, item.get('name',''), item.get('amt', 0),
                             item.get('supportBu',''), item.get('supportDept',''))
                        )

        # Trim to max_history
        old = conn.execute(
            "SELECT id FROM uploads ORDER BY uploaded_at DESC LIMIT -1 OFFSET ?",
            (max_history,)
        ).fetchall()
        for r in old:
            conn.execute("DELETE FROM uploads WHERE id=?", (r['id'],))

        conn.commit()

    return jsonify({'ok': True, 'upload_id': upload_id})

# ── History list ──────────────────────────────────────────────────────────────
@app.route('/api/history')
def history_api():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id,filename,bu_names,uploaded_at,case_count,conf_summary "
            "FROM uploads ORDER BY uploaded_at DESC"
        ).fetchall()
        return jsonify([{
            'id':          r['id'],
            'filename':    r['filename'],
            'bu_names':    r['bu_names'],
            'uploaded_at': r['uploaded_at'],
            'case_count':  r['case_count'],
            'conf_summary': json.loads(r['conf_summary']) if r['conf_summary'] else {}
        } for r in rows])

# ── Load historical data (returns same structure as frontend RAW) ─────────────
@app.route('/api/data/<int:upload_id>')
def data_api(upload_id):
    with get_db() as conn:
        upload = conn.execute("SELECT * FROM uploads WHERE id=?", (upload_id,)).fetchone()
        if not upload:
            abort(404)

        cases = conn.execute(
            "SELECT id,case_no,cust_name,case_name,rep,bu,close_date,dept,support_bu,support_dept "
            "FROM cases WHERE upload_id=? ORDER BY bu,rep,close_date",
            (upload_id,)
        ).fetchall()

        raw    = {}
        bu_set = []
        for cas in cases:
            bu = cas['bu']
            if bu not in raw:
                raw[bu] = []
                bu_set.append(bu)

            items = conn.execute(
                "SELECT conf,item_name,amount,support_bu,support_dept FROM case_items WHERE case_id=?",
                (cas['id'],)
            ).fetchall()

            conf_dict = {}
            for item in items:
                c = item['conf']
                if c not in conf_dict:
                    conf_dict[c] = []
                conf_dict[c].append({
                    'name':        item['item_name'],
                    'amt':         item['amount'],
                    'supportBu':   item['support_bu']   or '',
                    'supportDept': item['support_dept'] or ''
                })

            raw[bu].append({
                'no':   cas['case_no'],   'cust': cas['cust_name'],
                'case': cas['case_name'], 'rep':  cas['rep'],
                'date': cas['close_date'],'conf': conf_dict,
                'dept':        cas['dept']         or '',
                'supportBu':   cas['support_bu']   or '',
                'supportDept': cas['support_dept'] or ''
            })

        bu_names = upload['bu_names'].split(',') if upload['bu_names'] else bu_set

    return jsonify({
        'raw':        raw,
        'buOrder':    bu_names,
        'filename':   upload['filename'],
        'uploadedAt': upload['uploaded_at'],
        'caseCount':  upload['case_count']
    })

# ── Delete upload ─────────────────────────────────────────────────────────────
@app.route('/api/upload/<int:upload_id>', methods=['DELETE'])
def delete_upload(upload_id):
    with get_db() as conn:
        conn.execute("DELETE FROM uploads WHERE id=?", (upload_id,))
        conn.commit()
    return jsonify({'ok': True})

# ── Compare two uploads ───────────────────────────────────────────────────────
@app.route('/api/compare')
def compare_api():
    id1 = request.args.get('id1', type=int)
    id2 = request.args.get('id2', type=int)
    if not id1 or not id2:
        return jsonify({'error': 'Need id1 and id2'}), 400

    def get_cases(uid):
        with get_db() as conn:
            cases = conn.execute(
                "SELECT id,case_no,cust_name,case_name,rep,bu,close_date,dept FROM cases WHERE upload_id=?",
                (uid,)
            ).fetchall()
            result = {}
            for cas in cases:
                items = conn.execute(
                    "SELECT conf, SUM(amount) as total FROM case_items WHERE case_id=? GROUP BY conf",
                    (cas['id'],)
                ).fetchall()
                conf_totals = {i['conf']: i['total'] for i in items}
                result[cas['case_no']] = {
                    'cust': cas['cust_name'], 'case': cas['case_name'],
                    'rep':  cas['rep'],       'bu':   cas['bu'],
                    'close_date': cas['close_date'],
                    'conf_totals': conf_totals,
                    'total': sum(conf_totals.values())
                }
        return result

    def get_conf_summary(uid):
        with get_db() as conn:
            rows = conn.execute("""
                SELECT ci.conf, SUM(ci.amount) as total
                FROM case_items ci JOIN cases c ON c.id=ci.case_id
                WHERE c.upload_id=? GROUP BY ci.conf
            """, (uid,)).fetchall()
        return {r['conf']: r['total'] for r in rows}

    def get_info(uid):
        with get_db() as conn:
            r = conn.execute(
                "SELECT filename,uploaded_at,case_count FROM uploads WHERE id=?", (uid,)
            ).fetchone()
        return {'filename': r['filename'], 'uploaded_at': r['uploaded_at'],
                'case_count': r['case_count']} if r else {}

    cases1   = get_cases(id1)
    cases2   = get_cases(id2)
    summary1 = get_conf_summary(id1)
    summary2 = get_conf_summary(id2)

    CONF_PREF = ['Base', 'ACP', 'Option-補救', 'Option', 'Maximum']
    all_confs = set(summary1) | set(summary2)
    sorted_confs = [c for c in CONF_PREF if c in all_confs] + \
                   sorted(c for c in all_confs if c not in CONF_PREF)

    conf_delta = [{'conf': c,
                   'before': summary1.get(c, 0),
                   'after':  summary2.get(c, 0),
                   'delta':  summary2.get(c, 0) - summary1.get(c, 0)}
                  for c in sorted_confs]

    new_cases, removed_cases, changed_cases = [], [], []
    for no in set(cases1) | set(cases2):
        in1, in2 = no in cases1, no in cases2
        if in2 and not in1:
            new_cases.append({'no': no, **cases2[no]})
        elif in1 and not in2:
            removed_cases.append({'no': no, **cases1[no]})
        else:
            c1, c2 = cases1[no], cases2[no]
            all_c = set(c1['conf_totals']) | set(c2['conf_totals'])
            delta_confs = {}
            for c in all_c:
                a1, a2 = c1['conf_totals'].get(c, 0), c2['conf_totals'].get(c, 0)
                if abs(a1 - a2) > 0.01:
                    delta_confs[c] = {'before': a1, 'after': a2, 'delta': a2 - a1}

            # Primary conf = conf with the highest total amount
            def prim(conf_totals):
                nz = {k: v for k, v in conf_totals.items() if v > 0}
                return max(nz, key=lambda k: nz[k]) if nz else ''

            prim_before  = prim(c1['conf_totals'])
            prim_after   = prim(c2['conf_totals'])
            rep_changed  = c1['rep'] != c2['rep']
            amt_changed  = bool(delta_confs)
            date_changed = c1['close_date'] != c2['close_date']
            conf_changed = prim_before != prim_after

            if amt_changed or date_changed or conf_changed or rep_changed:
                changed_cases.append({
                    'no':  no,
                    'cust': c2['cust'], 'case': c2['case'],
                    'bu':   c2['bu'],   'rep':  c2['rep'],
                    'close_date_before': c1['close_date'],
                    'close_date_after':  c2['close_date'],
                    'prim_conf_before':  prim_before,
                    'prim_conf_after':   prim_after,
                    'rep_before':        c1['rep'],
                    'rep_after':         c2['rep'],
                    'total_before':      c1['total'],
                    'total_after':       c2['total'],
                    'total_delta':       c2['total'] - c1['total']
                })

    return jsonify({
        'info1': get_info(id1), 'info2': get_info(id2),
        'conf_delta':     conf_delta,
        'new_cases':      new_cases,
        'removed_cases':  removed_cases,
        'changed_cases':  changed_cases,
        'new_count':      len(new_cases),
        'removed_count':  len(removed_cases),
        'changed_count':  len(changed_cases)
    })

# ── Trend: all uploads time-series ────────────────────────────────────────────
@app.route('/api/trend')
def trend_api():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id,uploaded_at,conf_summary FROM uploads ORDER BY uploaded_at ASC"
        ).fetchall()
        return jsonify([{
            'id':          r['id'],
            'uploaded_at': r['uploaded_at'],
            'conf_summary': json.loads(r['conf_summary']) if r['conf_summary'] else {}
        } for r in rows])

# ── Achievement API (server-side computation) ─────────────────────────────────
@app.route('/api/achievement')
def achievement_api():
    try:
        year = int(request.args.get('year', datetime.now().year))
    except (TypeError, ValueError):
        year = datetime.now().year

    # 季度過濾：預設全選，可傳 quarters=1,2 等
    QUARTER_MONTHS = {
        1: ('01','02','03'),
        2: ('04','05','06'),
        3: ('07','08','09'),
        4: ('10','11','12'),
    }
    try:
        qs_param = request.args.get('quarters', '1,2,3,4')
        selected_qs = [int(q) for q in qs_param.split(',') if q.strip().isdigit() and 1 <= int(q.strip()) <= 4]
    except Exception:
        selected_qs = [1, 2, 3, 4]
    if not selected_qs:
        selected_qs = [1, 2, 3, 4]

    # 全季 → 不加月份過濾（相容舊行為）
    all_quarters = len(selected_qs) == 4
    months = []
    for q in selected_qs:
        months.extend(QUARTER_MONTHS[q])
    # close_date 格式為 YYYY/MM/DD，用 substr 取第 6-7 字元（月份）
    month_filter_sql = "" if all_quarters else \
        f"AND substr(c.close_date, 6, 2) IN ({','.join(['?']*len(months))})"
    month_params = [] if all_quarters else months

    with get_db() as conn:
        # Latest upload
        latest = conn.execute(
            "SELECT id FROM uploads ORDER BY uploaded_at DESC LIMIT 1"
        ).fetchone()

        # Quotas（年度 Quota，不受季度影響）
        quotas = conn.execute(
            "SELECT dim_type, dim_name, quota FROM quotas WHERE year=?", (year,)
        ).fetchall()

        # Orders：年度全數加入，不受季度過濾影響
        orders = conn.execute(
            "SELECT dim_type, dim_name, SUM(amount) as total FROM orders WHERE year=? GROUP BY dim_type, dim_name",
            (year,)
        ).fetchall()
        order_map = {}
        for r in orders:
            if r['dim_type'] not in order_map:
                order_map[r['dim_type']] = {}
            order_map[r['dim_type']][r['dim_name']] = r['total']

        # Forecast amounts：依季度過濾結案日期
        forecast = {'sales_bu': {}, 'sales_dept': {}, 'support_bu': {}, 'support_dept': {}}
        if latest:
            upload_id = latest['id']
            # sales_bu / sales_dept：以 case 層級欄位分組
            case_dim_queries = [
                ('sales_bu',   'c.bu',   "COALESCE(c.bu,'') != ''"),
                ('sales_dept', 'c.dept', "COALESCE(c.dept,'') != ''"),
            ]
            for dim_type, col_expr, cond in case_dim_queries:
                rows = conn.execute(f"""
                    SELECT {col_expr} AS dim_name, ci.conf, SUM(ci.amount) AS total
                    FROM cases c
                    JOIN case_items ci ON ci.case_id = c.id
                    WHERE c.upload_id = ? AND {cond} {month_filter_sql}
                    GROUP BY {col_expr}, ci.conf
                """, (upload_id, *month_params)).fetchall()
                dim_map = {}
                for r in rows:
                    n = r['dim_name']
                    if n not in dim_map:
                        dim_map[n] = {}
                    dim_map[n][r['conf']] = r['total']
                forecast[dim_type] = dim_map

            # support_bu / support_dept：以 case_items 層級欄位分組
            item_dim_queries = [
                ('support_bu',   'ci.support_bu',   "COALESCE(ci.support_bu,'') != ''"),
                ('support_dept', 'ci.support_dept', "COALESCE(ci.support_dept,'') != ''"),
            ]
            for dim_type, col_expr, cond in item_dim_queries:
                rows = conn.execute(f"""
                    SELECT {col_expr} AS dim_name, ci.conf, SUM(ci.amount) AS total
                    FROM cases c
                    JOIN case_items ci ON ci.case_id = c.id
                    WHERE c.upload_id = ? AND {cond} {month_filter_sql}
                    GROUP BY {col_expr}, ci.conf
                """, (upload_id, *month_params)).fetchall()
                dim_map = {}
                for r in rows:
                    n = r['dim_name']
                    if n not in dim_map:
                        dim_map[n] = {}
                    dim_map[n][r['conf']] = r['total']
                forecast[dim_type] = dim_map

        return jsonify({
            'quotas':    [{'dim_type': r['dim_type'], 'dim_name': r['dim_name'], 'quota': r['quota']} for r in quotas],
            'order_map': order_map,
            'forecast':  forecast
        })

# ── Quotas CRUD ───────────────────────────────────────────────────────────────
@app.route('/api/quotas', methods=['GET'])
def quotas_get():
    year = request.args.get('year', type=int)
    with get_db() as conn:
        q = "SELECT id,year,dim_type,dim_name,quota,updated_at FROM quotas"
        args = ()
        if year:
            q += " WHERE year=?"
            args = (year,)
        q += " ORDER BY year DESC, dim_type, dim_name"
        rows = conn.execute(q, args).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/quotas', methods=['POST'])
def quotas_post():
    d = request.get_json()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO quotas (year,dim_type,dim_name,quota,updated_at) VALUES (?,?,?,?,?)",
            (d['year'], d['dim_type'], d['dim_name'], d['quota'], now)
        )
        conn.commit()
    return jsonify({'ok': True, 'id': cur.lastrowid})

@app.route('/api/quotas/<int:qid>', methods=['PUT'])
def quotas_put(qid):
    d = request.get_json()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        conn.execute(
            "UPDATE quotas SET year=?,dim_type=?,dim_name=?,quota=?,updated_at=? WHERE id=?",
            (d['year'], d['dim_type'], d['dim_name'], d['quota'], now, qid)
        )
        conn.commit()
    return jsonify({'ok': True})

@app.route('/api/quotas/<int:qid>', methods=['DELETE'])
def quotas_delete(qid):
    with get_db() as conn:
        conn.execute("DELETE FROM quotas WHERE id=?", (qid,))
        conn.commit()
    return jsonify({'ok': True})

# ── Orders CRUD ────────────────────────────────────────────────────────────────
@app.route('/api/orders', methods=['GET'])
def orders_get():
    year = request.args.get('year', type=int)
    with get_db() as conn:
        q = "SELECT id,year,dim_type,dim_name,amount,source,updated_at FROM orders"
        args = ()
        if year:
            q += " WHERE year=?"
            args = (year,)
        q += " ORDER BY year DESC, dim_type, dim_name"
        rows = conn.execute(q, args).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/orders', methods=['POST'])
def orders_post():
    d = request.get_json()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO orders (year,dim_type,dim_name,amount,updated_at) VALUES (?,?,?,?,?)",
            (d['year'], d['dim_type'], d['dim_name'], d['amount'], now)
        )
        conn.commit()
    return jsonify({'ok': True, 'id': cur.lastrowid})

@app.route('/api/orders/<int:oid>', methods=['PUT'])
def orders_put(oid):
    d = request.get_json()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        conn.execute(
            "UPDATE orders SET year=?,dim_type=?,dim_name=?,amount=?,updated_at=? WHERE id=?",
            (d['year'], d['dim_type'], d['dim_name'], d['amount'], now, oid)
        )
        conn.commit()
    return jsonify({'ok': True})

@app.route('/api/orders/<int:oid>', methods=['DELETE'])
def orders_delete(oid):
    with get_db() as conn:
        conn.execute("DELETE FROM orders WHERE id=?", (oid,))
        conn.commit()
    return jsonify({'ok': True})

# ── Orders: source info ───────────────────────────────────────────────────────
@app.route('/api/order_uploads', methods=['GET'])
def order_uploads_list():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id,filename,uploaded_at,year,row_count FROM order_uploads ORDER BY uploaded_at DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

# ── Orders: delete upload record (cascade records + re-aggregate orders) ──────
@app.route('/api/order_uploads/<int:uid>', methods=['DELETE'])
def order_upload_delete(uid):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        # 找出這次上傳涵蓋的年度
        years = [r['register_year'] for r in conn.execute(
            "SELECT DISTINCT register_year FROM order_records WHERE upload_id=? AND register_year > 0",
            (uid,)
        ).fetchall()]
        # 刪除 order_uploads（CASCADE 自動清除 order_records）
        conn.execute("DELETE FROM order_uploads WHERE id=?", (uid,))
        # 針對每個年度：先清掉舊彙總，再從剩餘 records 重新計算
        dim_map = {
            'sales_bu':    'biz_bu',
            'sales_dept':  'biz_dept',
            'support_bu':  'support_bu',
            'support_dept':'support_dept',
        }
        for yr in years:
            conn.execute("DELETE FROM orders WHERE year=? AND source='upload'", (yr,))
            for dim_type, field in dim_map.items():
                rows = conn.execute(f"""
                    SELECT {field} AS dim_name, SUM(amount_pretax) AS total
                    FROM order_records
                    WHERE register_year=? AND {field} != ''
                    GROUP BY {field}
                """, (yr,)).fetchall()
                for row in rows:
                    conn.execute(
                        "INSERT INTO orders (year,dim_type,dim_name,amount,source,updated_at) VALUES (?,?,?,?,?,?)",
                        (yr, dim_type, row['dim_name'], row['total'], 'upload', now)
                    )
        conn.commit()
    return jsonify({'ok': True})

# ── Orders: upload XLS and parse ──────────────────────────────────────────────
@app.route('/api/orders/upload', methods=['POST'])
def orders_upload():
    if 'file' not in request.files:
        return jsonify({'error': '未收到檔案'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': '檔名為空'}), 400

    # ── 讀取 Excel（支援 .xls 與 .xlsx）──────────────────────────────
    file_bytes = f.read()
    fname_lower = f.filename.lower()

    if fname_lower.endswith('.xlsx'):
        try:
            import openpyxl
        except ImportError:
            return jsonify({'error': '伺服器缺少 openpyxl 套件，請執行 pip install openpyxl'}), 500
        try:
            wb_px = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws_px = wb_px.active
            _rows = [[c.value for c in row] for row in ws_px.iter_rows()]
            wb_px.close()
        except Exception as e:
            return jsonify({'error': f'無法開啟 Excel 檔案：{e}'}), 400

        def cell_str(row, col_idx):
            try:
                v = _rows[row][col_idx]
                return str(v).strip() if v is not None else ''
            except Exception:
                return ''

        def cell_date(row, c_name):
            if c_name not in col:
                return ''
            try:
                v = _rows[row][col[c_name]]
                if v is None:
                    return ''
                if hasattr(v, 'year'):           # datetime / date object
                    return f'{v.year:04d}/{v.month:02d}/{v.day:02d}'
                s = str(v).strip()
                if len(s) >= 8 and ('/' in s or '-' in s):
                    return s[:10].replace('-', '/')
                return ''
            except Exception:
                return ''

        nrows = len(_rows)
        ncols = len(_rows[0]) if _rows else 0

    else:  # .xls
        try:
            import xlrd
        except ImportError:
            return jsonify({'error': '伺服器缺少 xlrd 套件，請執行 pip install xlrd==1.2.0'}), 500
        try:
            wb_xl = xlrd.open_workbook(file_contents=file_bytes)
            ws_xl = wb_xl.sheets()[0]
        except Exception as e:
            return jsonify({'error': f'無法開啟 Excel 檔案：{e}'}), 400

        def cell_str(row, col_idx):
            try:
                v = ws_xl.cell_value(row, col_idx)
                return str(v).strip() if v is not None else ''
            except Exception:
                return ''

        def cell_date(row, c_name):
            if c_name not in col:
                return ''
            try:
                cell = ws_xl.cell(row, col[c_name])
                if cell.ctype == xlrd.XL_CELL_DATE:
                    t = xlrd.xldate_as_tuple(cell.value, wb_xl.datemode)
                    return f'{t[0]:04d}/{t[1]:02d}/{t[2]:02d}'
                v = str(cell.value).strip()
                if len(v) >= 8 and ('/' in v or '-' in v):
                    return v[:10].replace('-', '/')
                return ''
            except Exception:
                return ''

        nrows = ws_xl.nrows
        ncols = ws_xl.ncols

    # ── 偵測標題列 ────────────────────────────────────────────────────
    first_cell = cell_str(0, 0)
    if any(k in first_cell for k in ('CN', 'SFM', '每月', '合約', '訂單')):
        header_row = 1
        data_start = 2
    else:
        header_row = 0
        data_start = 1

    # 建立欄名→index 對照
    col = {}
    for ci in range(ncols):
        h = cell_str(header_row, ci)
        if h:
            col[h] = ci

    # 必要欄位檢查
    REQ = ['分項金額作帳幣別總額-未稅', '最新合約承辦BU', '支援BU']
    missing = [r for r in REQ if r not in col]
    if missing:
        return jsonify({'error': f'找不到必要欄位：{", ".join(missing)}'}), 400

    def cell_num(row, c_name):
        if c_name not in col:
            return 0.0
        try:
            if fname_lower.endswith('.xlsx'):
                v = _rows[row][col[c_name]]
            else:
                v = ws_xl.cell_value(row, col[c_name])
            return float(v) if v not in ('', None) else 0.0
        except Exception:
            return 0.0

    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    records = []

    for r in range(data_start, nrows):
        # 跳過空白列（以承辦BU 為主要判斷欄）
        biz_bu = cell_str(r, col.get('最新合約承辦BU', -1)) if '最新合約承辦BU' in col else ''
        if not biz_bu:
            continue

        reg_date = cell_date(r, '合約登記日')
        try:
            parts  = reg_date.split('/')
            r_year = int(parts[0])
            r_mon  = int(parts[1])
        except Exception:
            r_year, r_mon = 0, 0

        amt = cell_num(r, '分項金額作帳幣別總額-未稅')

        records.append({
            'contract_no':      cell_str(r, col.get('合約編號',    -1)),
            'contract_no_bu':   cell_str(r, col.get('合約編號(BU)', -1)),
            'contract_name':    cell_str(r, col.get('合約名稱',     -1)),
            'cust_name':        cell_str(r, col.get('客戶中文名稱', -1)) or cell_str(r, col.get('客戶名稱', -1)),
            'register_date':    reg_date,
            'register_year':    r_year,
            'register_month':   r_mon,
            'amount_pretax':    amt,
            'biz_bu':           biz_bu,
            'biz_dept':         cell_str(r, col.get('最新合約承辦部門', -1)),
            'biz_rep':          (cell_str(r, col.get('最新合約承辦人姓名', -1))
                                 or cell_str(r, col.get('業務代表',       -1))
                                 or cell_str(r, col.get('業務人員',       -1))),
            'support_bu':       cell_str(r, col.get('支援BU',           -1)),
            'support_dept':     cell_str(r, col.get('支援部門',          -1)),
            'transaction_type': cell_str(r, col.get('交易型態',          -1)),
            'contract_type':    cell_str(r, col.get('合約類別',          -1)),
            'industry':         cell_str(r, col.get('客戶行業別',        -1)),
            'industry_class':   cell_str(r, col.get('客戶行業別分類',    -1)),
            'product_line':     cell_str(r, col.get('產品線說明',        -1)),
            'biz_line':         cell_str(r, col.get('BIZ線說明',        -1)),
            'vendor':           cell_str(r, col.get('廠商',              -1)),
            'product_name':     cell_str(r, col.get('產品/專案名稱',     -1)),
            'item_seq':         cell_str(r, col.get('分項項次',          -1)),
            'case_no':          cell_str(r, col.get('銷售案編號',        -1)),
        })

    if not records:
        return jsonify({'error': '檔案中未找到有效資料列'}), 400

    # 取得所有年度（用於後續彙總）
    years = sorted({rec['register_year'] for rec in records if rec['register_year'] > 0})

    with get_db() as conn:
        # 新增 order_upload 記錄
        cur = conn.execute(
            "INSERT INTO order_uploads (filename,uploaded_at,year,row_count) VALUES (?,?,?,?)",
            (f.filename, now, years[0] if years else 0, len(records))
        )
        upload_id = cur.lastrowid

        # 存入 order_records
        for rec in records:
            conn.execute("""
                INSERT INTO order_records
                (upload_id,contract_no,contract_no_bu,contract_name,cust_name,
                 register_date,register_year,register_month,amount_pretax,
                 biz_bu,biz_dept,biz_rep,support_bu,support_dept,
                 transaction_type,contract_type,industry,industry_class,
                 product_line,biz_line,vendor,product_name,item_seq,case_no)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                upload_id,
                rec['contract_no'], rec['contract_no_bu'], rec['contract_name'], rec['cust_name'],
                rec['register_date'], rec['register_year'], rec['register_month'], rec['amount_pretax'],
                rec['biz_bu'], rec['biz_dept'], rec['biz_rep'], rec['support_bu'], rec['support_dept'],
                rec['transaction_type'], rec['contract_type'], rec['industry'], rec['industry_class'],
                rec['product_line'], rec['biz_line'], rec['vendor'], rec['product_name'],
                rec['item_seq'], rec['case_no']
            ))

        # 依年度自動彙總，更新 orders 表（取代 source='upload' 的舊資料）
        for yr in years:
            # 先刪除該年度的上傳來源訂單
            conn.execute("DELETE FROM orders WHERE year=? AND source='upload'", (yr,))

            dim_map = {
                'sales_bu':    ('biz_bu',      '最新合約承辦BU'),
                'sales_dept':  ('biz_dept',    '最新合約承辦部門'),
                'support_bu':  ('support_bu',  '支援BU'),
                'support_dept':('support_dept','支援部門'),
            }
            for dim_type, (field, _) in dim_map.items():
                rows = conn.execute(f"""
                    SELECT {field} AS dim_name, SUM(amount_pretax) AS total
                    FROM order_records
                    WHERE upload_id=? AND register_year=? AND {field} != ''
                    GROUP BY {field}
                """, (upload_id, yr)).fetchall()
                for row in rows:
                    conn.execute(
                        "INSERT INTO orders (year,dim_type,dim_name,amount,source,updated_at) VALUES (?,?,?,?,?,?)",
                        (yr, dim_type, row['dim_name'], row['total'], 'upload', now)
                    )

        conn.commit()

    return jsonify({
        'ok':        True,
        'upload_id': upload_id,
        'row_count': len(records),
        'years':     years,
        'filename':  f.filename
    })

# ── Orders: detail records ─────────────────────────────────────────────────────
@app.route('/api/orders/detail', methods=['GET'])
def orders_detail():
    year  = request.args.get('year',  type=int)
    month = request.args.get('month', type=int)   # 0 = all
    dim   = request.args.get('dim',   default='support_bu')   # support_bu / support_dept

    # 取最新一次的 order_upload
    with get_db() as conn:
        latest_upload = conn.execute(
            "SELECT id FROM order_uploads ORDER BY uploaded_at DESC LIMIT 1"
        ).fetchone()
        if not latest_upload:
            return jsonify({'records': [], 'monthly_totals': []})

        uid = latest_upload['id']
        conditions = ["upload_id=?"]
        params: list = [uid]
        if year:
            conditions.append("register_year=?")
            params.append(year)
        if month:
            conditions.append("register_month=?")
            params.append(month)

        where = " AND ".join(conditions)
        rows = conn.execute(f"""
            SELECT contract_no, contract_no_bu, contract_name, cust_name,
                   register_date, register_year, register_month,
                   amount_pretax, biz_bu, biz_dept, biz_rep, support_bu, support_dept,
                   transaction_type, contract_type, industry, product_line,
                   vendor, product_name, item_seq, case_no
            FROM order_records WHERE {where}
            ORDER BY register_year DESC, register_month DESC, support_bu, support_dept, contract_name
        """, params).fetchall()

        # 月份加總（依維度 + 月份）
        dim_field = 'support_bu' if dim == 'support_bu' else 'support_dept'
        monthly = conn.execute(f"""
            SELECT register_year, register_month, {dim_field} AS dim_name,
                   SUM(amount_pretax) AS total, COUNT(*) AS cnt
            FROM order_records WHERE {where}
            AND {dim_field} != ''
            GROUP BY register_year, register_month, {dim_field}
            ORDER BY register_year DESC, register_month DESC, {dim_field}
        """, params).fetchall()

    return jsonify({
        'records':       [dict(r) for r in rows],
        'monthly_totals': [dict(r) for r in monthly]
    })

# ── Orders: analysis stats ─────────────────────────────────────────────────────
@app.route('/api/orders/analysis', methods=['GET'])
def orders_analysis():
    year = request.args.get('year', type=int, default=datetime.now().year)

    with get_db() as conn:
        latest_upload = conn.execute(
            "SELECT id FROM order_uploads ORDER BY uploaded_at DESC LIMIT 1"
        ).fetchone()
        if not latest_upload:
            return jsonify({})

        uid = latest_upload['id']
        base_cond = "upload_id=? AND register_year=?"
        bp = (uid, year)

        def agg(field):
            rows = conn.execute(f"""
                SELECT {field} AS name, SUM(amount_pretax) AS total, COUNT(*) AS cnt
                FROM order_records WHERE {base_cond} AND {field} != ''
                GROUP BY {field} ORDER BY total DESC
            """, bp).fetchall()
            return [{'name': r['name'], 'total': r['total'], 'cnt': r['cnt']} for r in rows]

        # 月度趨勢
        monthly = conn.execute(f"""
            SELECT register_month AS month, SUM(amount_pretax) AS total, COUNT(*) AS cnt
            FROM order_records WHERE {base_cond}
            GROUP BY register_month ORDER BY register_month
        """, bp).fetchall()

        # Grand total
        grand = conn.execute(f"""
            SELECT SUM(amount_pretax) AS total, COUNT(*) AS cnt
            FROM order_records WHERE {base_cond}
        """, bp).fetchone()

    return jsonify({
        'year':             year,
        'grand_total':      grand['total'] or 0,
        'grand_count':      grand['cnt']   or 0,
        'transaction_type': agg('transaction_type'),
        'contract_type':    agg('contract_type'),
        'industry':         agg('industry'),
        'industry_class':   agg('industry_class'),
        'product_line':     agg('product_line'),
        'biz_line':         agg('biz_line'),
        'support_bu':       agg('support_bu'),
        'biz_bu':           agg('biz_bu'),
        'monthly':          [{'month': r['month'], 'total': r['total'], 'cnt': r['cnt']} for r in monthly],
    })

# ── Opportunity Tracking ──────────────────────────────────────────────────────

@app.route('/api/opp_tracking/summary', methods=['GET'])
def opp_tracking_summary():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT case_no, COUNT(*) AS total, SUM(is_done) AS done FROM opp_tracking GROUP BY case_no"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/opp_tracking', methods=['GET'])
def opp_tracking_get():
    case_no = request.args.get('case_no', '')
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, case_no, case_name, cust_name, note, is_done, done_note, done_at, created_at FROM opp_tracking WHERE case_no=? ORDER BY id ASC",
            (case_no,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/opp_tracking', methods=['POST'])
def opp_tracking_post():
    data = request.get_json() or {}
    case_no   = data.get('case_no', '').strip()
    case_name = data.get('case_name', '').strip()
    cust_name = data.get('cust_name', '').strip()
    note      = data.get('note', '').strip()
    if not case_no or not note:
        return jsonify({'error': '缺少必要欄位'}), 400
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO opp_tracking (case_no, case_name, cust_name, note, is_done, created_at) VALUES (?,?,?,?,0,?)",
            (case_no, case_name, cust_name, note, now)
        )
        conn.commit()
        row = conn.execute("SELECT id, case_no, case_name, cust_name, note, is_done, done_note, done_at, created_at FROM opp_tracking WHERE id=?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row))

@app.route('/api/opp_tracking/<int:tid>', methods=['PUT'])
def opp_tracking_put(tid):
    data = request.get_json() or {}
    # Inline note edit: only update note field
    if 'note' in data:
        note = str(data['note']).strip()
        if not note:
            return jsonify({'ok': False, 'error': 'note cannot be empty'}), 400
        with get_db() as conn:
            conn.execute("UPDATE opp_tracking SET note=? WHERE id=?", (note, tid))
            conn.commit()
        return jsonify({'ok': True})
    is_done   = int(bool(data.get('is_done', False)))
    done_note = data.get('done_note', None)
    done_at   = data.get('done_at',   None)
    with get_db() as conn:
        if done_note is not None or done_at is not None:
            conn.execute(
                "UPDATE opp_tracking SET is_done=?, done_note=?, done_at=? WHERE id=?",
                (is_done, done_note or '', done_at or '', tid)
            )
        else:
            conn.execute("UPDATE opp_tracking SET is_done=? WHERE id=?", (is_done, tid))
        conn.commit()
    return jsonify({'ok': True})

@app.route('/api/opp_tracking/<int:tid>', methods=['DELETE'])
def opp_tracking_delete(tid):
    with get_db() as conn:
        conn.execute("DELETE FROM opp_tracking WHERE id=?", (tid,))
        conn.commit()
    return jsonify({'ok': True})

@app.route('/api/opp_tracking/all', methods=['GET'])
def opp_tracking_all():
    filter_done = request.args.get('done', default=None)
    with get_db() as conn:
        if filter_done is None:
            rows = conn.execute(
                "SELECT id, case_no, case_name, cust_name, note, is_done, done_note, done_at, created_at FROM opp_tracking ORDER BY is_done ASC, id DESC"
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, case_no, case_name, cust_name, note, is_done, done_note, done_at, created_at FROM opp_tracking WHERE is_done=? ORDER BY id DESC",
                (int(filter_done),)
            ).fetchall()
    return jsonify([dict(r) for r in rows])

# ── Main ──────────────────────────────────────────────────────────────────────

def open_browser():
    time.sleep(1.3)
    webbrowser.open('http://localhost:8000')

if __name__ == '__main__':
    init_db()
    threading.Thread(target=open_browser, daemon=True).start()
    print("=" * 52)
    print("  銷售案 Forecast 分析系統  已啟動")
    print("  網址：http://localhost:8000")
    print("  按 Ctrl+C 可停止服務")
    print("=" * 52)
    app.run(host='127.0.0.1', port=8000, debug=False)
